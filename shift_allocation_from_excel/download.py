import frappe
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from frappe.utils import getdate, add_days, now_datetime
import io

@frappe.whitelist()
def download_shift_template(start_date, end_date, docname):
    start_date = getdate(start_date)
    end_date = getdate(end_date)
    total_days = (end_date - start_date).days + 1

    doc = frappe.get_doc("Shift Allocation Tool", docname)
    employees = doc.get("shift_allocation_employees") or []

    shift_types = [d.name for d in frappe.get_all("Shift Type")]
    if "WO" not in shift_types:
        shift_types.append("WO")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shift Template"

    hidden_sheet = wb.create_sheet(title="DropdownData")
    for i, st in enumerate(shift_types, start=1):
        hidden_sheet.cell(row=i, column=1, value=st)
    hidden_sheet.sheet_state = 'hidden'
    dv_range_name = "ShiftTypes"
    wb.create_named_range(dv_range_name, hidden_sheet, f"$A$1:$A${len(shift_types)}")

    headers = ["S. No", "EMPLOYEE ID", "EMPLOYEE NAME"]
    date_list = []

    for i in range(total_days):
        current_date = add_days(start_date, i)
        headers.append(current_date.strftime("%-d-%b-%Y"))
        date_list.append(current_date)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    for i, current_date in enumerate(date_list):
        weekday = current_date.strftime("%A")
        col_idx = i + 4
        cell = ws.cell(row=2, column=col_idx, value=weekday)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(italic=True)
        if employees:
            dv = DataValidation(type="list", formula1=f"={dv_range_name}", allow_blank=True)
            ws.add_data_validation(dv)
            last_row = len(employees) + 2 
            dv.add(f"{get_column_letter(col_idx)}3:{get_column_letter(col_idx)}{last_row}")
        #dv = DataValidation(type="list", formula1=f"={dv_range_name}", allow_blank=True)
        #ws.add_data_validation(dv)
        #last_row = len(employees) + 2 
        #dv.add(f"{get_column_letter(col_idx)}3:{get_column_letter(col_idx)}{last_row}")
    for idx, emp in enumerate(employees, start=1):
        row = idx + 2
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=emp.employee)
        ws.cell(row=row, column=3, value=emp.employee_name)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    frappe.response["filename"] = f"Shift_Template_{start_date}_to_{end_date}_current_{now_datetime().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    frappe.response["filecontent"] = buffer.read()
    frappe.response["type"] = "binary"

