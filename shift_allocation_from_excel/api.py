import frappe
import openpyxl
from frappe.utils import getdate, add_days
from datetime import datetime
from collections import defaultdict

def add_wo_to_holiday_list(start_date, end_date, employee_id, holiday_list_name):
    holiday_list = frappe.get_doc("Holiday List", holiday_list_name)
    existing_dates = {h.holiday_date for h in holiday_list.holidays}

    date = getdate(start_date)
    end = getdate(end_date)
    added = False

    while date <= end:
        if date not in existing_dates:
            day_name = date.strftime("%A")
            holiday_list.append("holidays", {
                "holiday_date": date,
                "weekly_off": 1,
                "description": f"{day_name} Weekly Off"
            })
            added = True
        date = add_days(date, 1)

    if added:
        holiday_list.save()

def process_shift_excel(doc, method):
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.shift})
        file_path = file_doc.get_full_path()
    except Exception:
        frappe.log_error(frappe.get_traceback(), "❌ Failed to retrieve Excel file")
        frappe.throw("Could not retrieve or read the attached file.")

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    processed, skipped = 0, 0
    date_headers = []

    for col in range(4, sheet.max_column + 1):
        val = sheet.cell(row=1, column=col).value
        if not val:
            continue
        try:
            date_headers.append(getdate(val))
        except Exception:
            frappe.log_error(f"Invalid date format in header: {val}", "Shift Upload")
            frappe.throw(f"Invalid date header: {val}")

    employee_holiday_map = {}
    shift_assignments = []

    for row in range(3, sheet.max_row + 1):
        barcode_cell = sheet.cell(row=row, column=2).value
        if not barcode_cell:
            skipped += 1
            continue

        barcode = str(barcode_cell).strip()
        employee_name = frappe.db.get_value("Employee", barcode)
        if not employee_name:
            frappe.log_error(f"Employee with barcode {barcode} not found.", "Shift Upload")
            skipped += 1
            continue

        if employee_name not in employee_holiday_map:
            holiday_list_name = frappe.db.get_value("Employee", employee_name, "holiday_list")
            if not holiday_list_name:
                frappe.log_error(f"Employee {employee_name} has no Holiday List.", "Shift Upload")
                skipped += 1
                continue
            employee_holiday_map[employee_name] = holiday_list_name
        else:
            holiday_list_name = employee_holiday_map[employee_name]

        current_shift = None
        current_start = None
        current_end = None

        for i, date in enumerate(date_headers):
            shift_val = sheet.cell(row=row, column=i + 4).value
            shift_code = str(shift_val).strip().upper() if shift_val else ""

            if shift_code == current_shift:
                current_end = date
            else:
                if current_shift:
                    try:
                        if current_shift and current_shift != "WO":
                            shift_assignments.append({
                                "employee": employee_name,
                                "shift_type": current_shift,
                                "start_date": current_start,
                                "end_date": current_end
                            })
                            processed += 1
                        elif current_shift == "WO":
                            add_wo_to_holiday_list(current_start, current_end, employee_name, holiday_list_name)
                    except Exception:
                        frappe.log_error(frappe.get_traceback(), f"❌ Error for employee {employee_name}")
                        skipped += 1

                current_shift = shift_code
                current_start = date
                current_end = date

        if current_shift:
            try:
                if current_shift and current_shift != "WO":
                    shift_assignments.append({
                        "employee": employee_name,
                        "shift_type": current_shift,
                        "start_date": current_start,
                        "end_date": current_end
                    })
                    processed += 1
                elif current_shift == "WO":
                    add_wo_to_holiday_list(current_start, current_end, employee_name, holiday_list_name)
            except Exception:
                frappe.log_error(frappe.get_traceback(), f"❌ Final shift error for employee {employee_name}")
                skipped += 1

    for data in shift_assignments:
        try:
            doc = frappe.get_doc({
                "doctype": "Shift Assignment",
                "employee": data["employee"],
                "shift_type": data["shift_type"],
                "start_date": data["start_date"],
                "end_date": data["end_date"]
            })
            doc.insert(ignore_permissions=True)
            doc.submit()
        except Exception as e:
            frappe.log_error(frappe.get_traceback(), f"❌ Failed to insert Shift Assignment for {data['employee']}")
            skipped += 1
            processed -= 1

    frappe.msgprint(f"✅ Shift assignment completed.<br>Processed: {processed}<br>Skipped: {skipped}")


@frappe.whitelist()
def get_employees(branch=None, department=None, designation=None,employee_grade=None, status="Active"):
    filters = {"status": status}
    if branch:
        filters["branch"] = branch
    if employee_grade:
        filters["grade"] = employee_grade

    if department:
        filters["department"] = department
    if designation:
        filters["designation"] = designation

    employees = frappe.get_all("Employee",
        filters=filters,
        fields=["name", "employee_name", "department","designation"]
    )

    return employees

